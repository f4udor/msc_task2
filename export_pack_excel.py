from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from packaging_pdf_parser import (
    EXCEL_FIELDS,
    FIELD_MODES,
    IMAGES_DIR,
    build_coverage_report,
    build_structured_record,
    iter_pdfs,
    normalize_field_value,
)


SOURCE_COLORS = {
    "static": "D9EAD3",
    "filename": "CFE2F3",
    "pdf_text": "FFF2CC",
    "ocr": "FCE5CD",
    "inference": "EAD1DC",
    "default_false": "EFEFEF",
    "missing": "F4CCCC",
}


def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def build_records_sheet(workbook: Workbook, records: list[dict[str, object]]) -> None:
    ws = workbook.active
    ws.title = "Records"

    headers = ["File", "Missing Fields", "Pages"]
    for _, field_name, label in EXCEL_FIELDS:
        headers.extend([label, f"{field_name}_source", f"{field_name}_confidence"])

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        row = [
            record["file"],
            record["missing_fields_count"],
            record["analysis"]["pages"],
        ]
        for _, field_name, _ in EXCEL_FIELDS:
            field = record["fields"][field_name]
            row.extend([field["value"], field["source"], field["confidence"]])
        ws.append(row)

    # Color value/source/confidence triplets by source.
    start_col = 4
    for row_idx in range(2, ws.max_row + 1):
        for field_offset, (_, _, _) in enumerate(EXCEL_FIELDS):
            field_col = start_col + field_offset * 3
            source = ws.cell(row=row_idx, column=field_col + 1).value
            fill_color = SOURCE_COLORS.get(str(source), "FFFFFF")
            fill = PatternFill(fill_type="solid", fgColor=fill_color)
            for col in range(field_col, field_col + 3):
                ws.cell(row=row_idx, column=col).fill = fill

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def build_normalized_sheet(workbook: Workbook, records: list[dict[str, object]]) -> None:
    ws = workbook.create_sheet("Normalized")
    headers = ["File", "Missing Fields", "Pages"]
    for _, _, label in EXCEL_FIELDS:
        headers.append(label)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        row = [
            record["file"],
            record["missing_fields_count"],
            record["analysis"]["pages"],
        ]
        for _, field_name, _ in EXCEL_FIELDS:
            row.append(normalize_field_value(record["fields"][field_name]))
        ws.append(row)

    # color cells by semantic mode instead of extraction source
    start_col = 4
    for field_offset, (_, field_name, _) in enumerate(EXCEL_FIELDS):
        fill_color = "E2F0D9" if FIELD_MODES[field_name] == "presence_only" else "DDEBF7"
        fill = PatternFill(fill_type="solid", fgColor=fill_color)
        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row=row_idx, column=start_col + field_offset).fill = fill

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def build_coverage_sheet(workbook: Workbook, coverage_report: dict[str, object]) -> None:
    ws = workbook.create_sheet("Coverage")
    headers = [
        "Column",
        "Field Name",
        "Label",
        "Mode",
        "Filled",
        "Total",
        "Coverage %",
        "Static",
        "Filename",
        "PDF Text",
        "OCR",
        "Inference",
        "Default False",
        "Missing",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for field in coverage_report["fields"]:
        source_counts = field["source_counts"]
        ws.append(
            [
                field["column"],
                field["field_name"],
                field["label"],
                FIELD_MODES[field["field_name"]],
                field["filled"],
                field["total"],
                field["coverage_pct"],
                source_counts["static"],
                source_counts["filename"],
                source_counts["pdf_text"],
                source_counts["ocr"],
                source_counts["inference"],
                source_counts["default_false"],
                source_counts["missing"],
            ]
        )

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def build_summary_sheet(workbook: Workbook, coverage_report: dict[str, object]) -> None:
    ws = workbook.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    ws.append(["Total PDFs", coverage_report["total_pdfs"]])
    ws.append(["Average Missing Fields", coverage_report["average_missing_fields"]])
    ws.append([])
    ws.append(["Best PDFs", "Missing Fields"])
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)
    for item in coverage_report["best_pdfs"]:
        ws.append([item["file"], item["missing_fields_count"]])

    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="Worst PDFs").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Missing Fields").font = Font(bold=True)
    for item in coverage_report["worst_pdfs"]:
        ws.append([item["file"], item["missing_fields_count"]])

    autofit_columns(ws)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Esporta in Excel i dati estratti dai PDF packaging, con dettaglio "
            "per campo e report di copertura."
        )
    )
    parser.add_argument(
        "directory",
        nargs="?",
        type=Path,
        default=IMAGES_DIR,
        help="Cartella con i PDF da analizzare. Default: ./images",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("pack_extraction_report.xlsx"),
        help="Percorso del file Excel da generare.",
    )
    parser.add_argument(
        "--ocr-dpi",
        type=int,
        default=200,
        help="Risoluzione OCR. Default: 200",
    )
    args = parser.parse_args()

    pdfs = iter_pdfs(args.directory.resolve())
    records = [
        build_structured_record(pdf_path, ocr_enabled=True, ocr_dpi=args.ocr_dpi)
        for pdf_path in pdfs
    ]
    coverage_report = build_coverage_report(records)

    workbook = Workbook()
    build_records_sheet(workbook, records)
    build_normalized_sheet(workbook, records)
    build_coverage_sheet(workbook, coverage_report)
    build_summary_sheet(workbook, coverage_report)

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
