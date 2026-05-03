from __future__ import annotations

import argparse
import statistics
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from packaging_pdf_parser import EXCEL_FIELDS, FIELD_MODES, IMAGES_DIR, build_coverage_report, build_structured_record, iter_pdfs, normalize_field_value


SOURCE_COLORS = {
    "static": "D9EAD3",
    "filename": "CFE2F3",
    "pdf_text": "FFF2CC",
    "ocr": "FCE5CD",
    "inference": "EAD1DC",
    "missing": "F4CCCC",
    "error": "F4CCCC",
}

STATUS_COLORS = {
    "perfect": "D9EAD3",
    "partial": "FFF2CC",
    "failed": "F4CCCC",
}

def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 48)


def summarize_dict(data: dict[str, int]) -> str:
    parts = [f"{key}={value}" for key, value in sorted(data.items()) if value]
    return ", ".join(parts)


def count_filled_sheet_values(sheet_row: dict[str, object]) -> int:
    return sum(1 for value in sheet_row.values() if value not in (None, ""))


def classify_record_status(record: dict[str, object]) -> str:
    if record.get("error"):
        return "failed"
    if record["missing_fields_count"] == 0 and not record["review"]["review_needed"]:
        return "perfect"
    return "partial"


def build_failed_record(pdf_path: Path, error: Exception) -> dict[str, object]:
    empty_row = {label: None for _, _, label in EXCEL_FIELDS}
    return {
        "file": pdf_path.name,
        "analysis": {
            "pages": 0,
            "vector_objects": 0,
            "raster_images": 0,
            "has_vector_content": False,
            "has_selectable_text": False,
            "word_count": 0,
        },
        "fields": {},
        "extra_fields": {},
        "anchors": {},
        "zones": {},
        "review": {
            "review_needed": True,
            "review_fields_count": len(EXCEL_FIELDS),
            "source_counts": {"error": len(EXCEL_FIELDS)},
            "confidence_counts": {"low": len(EXCEL_FIELDS)},
            "review_fields": [],
        },
        "sheet_row": empty_row,
        "missing_fields_count": len(EXCEL_FIELDS),
        "error": str(error),
    }


def collect_records(pdf_paths: list[Path], *, ocr_dpi: int) -> tuple[list[dict[str, object]], float]:
    records: list[dict[str, object]] = []
    started = time.perf_counter()

    for pdf_path in pdf_paths:
        per_file_started = time.perf_counter()
        try:
            record = build_structured_record(pdf_path, ocr_enabled=True, ocr_dpi=ocr_dpi)
        except Exception as exc:
            record = build_failed_record(pdf_path, exc)
        record["timing_seconds"] = round(time.perf_counter() - per_file_started, 3)
        record["status"] = classify_record_status(record)
        records.append(record)

    total_seconds = round(time.perf_counter() - started, 3)
    return records, total_seconds


def build_data_workbook(records: list[dict[str, object]], output_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Data"

    headers = ["File", "Status", "Review Needed", "Missing Fields", "Review Summary"]
    headers.extend(label for _, _, label in EXCEL_FIELDS)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        row = [
            record["file"],
            record["status"],
            record["review"]["review_needed"],
            record["missing_fields_count"],
            f"{record['review']['review_fields_count']} campi da verificare" if record["review"]["review_needed"] else "nessuna review",
        ]
        row.extend(record["sheet_row"][label] for _, _, label in EXCEL_FIELDS)
        ws.append(row)

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=2).value)
        fill = PatternFill(fill_type="solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = fill

    ws.freeze_panes = "A2"
    autofit_columns(ws)

    trace = workbook.create_sheet("Trace")
    trace_headers = ["File", "Status", "Field", "Value", "Source", "Confidence", "Issue Type", "Issue Reason"]
    trace.append(trace_headers)
    for cell in trace[1]:
        cell.font = Font(bold=True)

    for record in records:
        for _, field_name, label in EXCEL_FIELDS:
            field = record["fields"].get(field_name)
            if field is None:
                trace.append([record["file"], record["status"], label, None, "error", "low", "parser_error", "Errore di parsing del record."])
                continue
            review_match = next((item for item in record["review"]["review_fields"] if item["field_name"] == field_name), None)
            trace.append([
                record["file"],
                record["status"],
                label,
                normalize_field_value(field),
                field["source"],
                field["confidence"],
                review_match["issue_type"] if review_match else None,
                review_match["issue_reason"] if review_match else None,
            ])

    for row_idx in range(2, trace.max_row + 1):
        source = str(trace.cell(row=row_idx, column=5).value)
        fill = PatternFill(fill_type="solid", fgColor=SOURCE_COLORS.get(source, "FFFFFF"))
        for col in range(3, 9):
            trace.cell(row=row_idx, column=col).fill = fill

    trace.freeze_panes = "A2"
    autofit_columns(trace)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_log_workbook(
    records: list[dict[str, object]],
    coverage_report: dict[str, object],
    *,
    input_dir: Path,
    total_seconds: float,
    output_path: Path,
    ocr_dpi: int,
) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Run Summary"
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)

    status_counts = {
        "perfect": sum(1 for record in records if record["status"] == "perfect"),
        "partial": sum(1 for record in records if record["status"] == "partial"),
        "failed": sum(1 for record in records if record["status"] == "failed"),
    }
    timings = [record["timing_seconds"] for record in records]

    summary_rows = [
        ("Run UTC", datetime.now(timezone.utc).isoformat()),
        ("Input Directory", str(input_dir.resolve())),
        ("PDF Analizzati", len(records)),
        ("OCR Enabled", True),
        ("OCR DPI", ocr_dpi),
        ("Tempo Totale (s)", total_seconds),
        ("Tempo Medio per PDF (s)", round(statistics.mean(timings), 3) if timings else 0.0),
        ("Perfect", status_counts["perfect"]),
        ("Partial", status_counts["partial"]),
        ("Failed", status_counts["failed"]),
        ("Average Missing Fields", coverage_report["average_missing_fields"]),
    ]
    for row in summary_rows:
        summary.append(list(row))

    autofit_columns(summary)

    files_ws = workbook.create_sheet("File Log")
    file_headers = [
        "File",
        "Status",
        "Timing Seconds",
        "Pages",
        "Missing Fields",
        "Review Needed",
        "Review Fields Count",
        "Filled Fields",
        "Source Summary",
        "Confidence Summary",
        "Error",
        "Review Details",
    ]
    files_ws.append(file_headers)
    for cell in files_ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        files_ws.append([
            record["file"],
            record["status"],
            record["timing_seconds"],
            record["analysis"]["pages"],
            record["missing_fields_count"],
            record["review"]["review_needed"],
            record["review"]["review_fields_count"],
            count_filled_sheet_values(record["sheet_row"]),
            summarize_dict(record["review"]["source_counts"]),
            summarize_dict(record["review"]["confidence_counts"]),
            record.get("error"),
            " | ".join(
                f"{item['column']}:{item['issue_type']}"
                for item in record["review"]["review_fields"][:10]
            ),
        ])

    for row_idx in range(2, files_ws.max_row + 1):
        status = str(files_ws.cell(row=row_idx, column=2).value)
        fill = PatternFill(fill_type="solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        for col in range(1, 12):
            files_ws.cell(row=row_idx, column=col).fill = fill

    files_ws.freeze_panes = "A2"
    autofit_columns(files_ws)

    review_ws = workbook.create_sheet("Review Queue")
    review_headers = ["File", "Status", "Field", "Value", "Source", "Confidence", "Issue Type", "Issue Reason"]
    review_ws.append(review_headers)
    for cell in review_ws[1]:
        cell.font = Font(bold=True)

    for record in records:
        if record.get("error"):
            review_ws.append([record["file"], "failed", "record_error", record["error"], "error", "low", "parser_error", "Errore durante il parsing del PDF."])
            continue
        for item in record["review"]["review_fields"]:
            review_ws.append([
                record["file"],
                record["status"],
                item["label"],
                item["value"],
                item["source"],
                item["confidence"],
                item["issue_type"],
                item["issue_reason"],
            ])

    for row_idx in range(2, review_ws.max_row + 1):
        source = str(review_ws.cell(row=row_idx, column=5).value)
        fill = PatternFill(fill_type="solid", fgColor=SOURCE_COLORS.get(source, "FFFFFF"))
        for col in range(1, 9):
            review_ws.cell(row=row_idx, column=col).fill = fill

    review_ws.freeze_panes = "A2"
    autofit_columns(review_ws)

    coverage_ws = workbook.create_sheet("Coverage")
    coverage_headers = [
        "Column",
        "Field Name",
        "Label",
        "Mode",
        "Filled",
        "Total",
        "Coverage %",
        "Static",
        "Filename",
        "PDF Text",
        "OCR",
        "Inference",
        "Default False",
        "Missing",
    ]
    coverage_ws.append(coverage_headers)
    for cell in coverage_ws[1]:
        cell.font = Font(bold=True)

    for field in coverage_report["fields"]:
        source_counts = field["source_counts"]
        coverage_ws.append([
            field["column"],
            field["field_name"],
            field["label"],
            FIELD_MODES[field["field_name"]],
            field["filled"],
            field["total"],
            field["coverage_pct"],
            source_counts["static"],
            source_counts["filename"],
            source_counts["pdf_text"],
            source_counts["ocr"],
            source_counts["inference"],
            source_counts["default_false"],
            source_counts["missing"],
        ])

    coverage_ws.freeze_panes = "A2"
    autofit_columns(coverage_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Workflow finale packaging: analizza tutti i PDF in una cartella, "
            "genera un Excel dati e un Excel log con errori, review e copertura."
        )
    )
    parser.add_argument(
        "directory",
        nargs="?",
        type=Path,
        default=IMAGES_DIR,
        help="Cartella con i PDF da processare. Default: ./images",
    )
    parser.add_argument(
        "--data-output",
        type=Path,
        default=Path("pack_data.xlsx"),
        help="Excel finale con i dati raccolti.",
    )
    parser.add_argument(
        "--log-output",
        type=Path,
        default=Path("pack_log.xlsx"),
        help="Excel finale con log, errori e review.",
    )
    parser.add_argument(
        "--ocr-dpi",
        type=int,
        default=200,
        help="Risoluzione OCR. Default: 200",
    )
    args = parser.parse_args()

    input_dir = args.directory.resolve()
    if not input_dir.is_dir():
        print(f"Cartella non trovata: {input_dir}")
        return 1

    pdf_paths = iter_pdfs(input_dir)
    if not pdf_paths:
        print(f"Nessun PDF trovato in: {input_dir}")
        return 0

    records, total_seconds = collect_records(pdf_paths, ocr_dpi=args.ocr_dpi)
    successful_records = [record for record in records if not record.get("error")]
    coverage_report = build_coverage_report(successful_records) if successful_records else {
        "total_pdfs": 0,
        "average_missing_fields": 0.0,
        "best_pdfs": [],
        "worst_pdfs": [],
        "fields": [],
    }

    build_data_workbook(records, args.data_output.resolve())
    build_log_workbook(
        records,
        coverage_report,
        input_dir=input_dir,
        total_seconds=total_seconds,
        output_path=args.log_output.resolve(),
        ocr_dpi=args.ocr_dpi,
    )

    print(args.data_output.resolve())
    print(args.log_output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
